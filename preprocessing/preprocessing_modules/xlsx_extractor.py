from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from typing import List, Dict, Any
from PIL import Image
from io import BytesIO
import pytesseract
import os
from logger.custom_logger import CustomLogger

# module logger
logger = CustomLogger().get_logger(__file__)

def extract_xlsx(xlsx_path: str, tesseract_cmd: str = None) -> List[Dict[str, Any]]:
    """
    Extracts content from an XLSX file, including table data and OCR'd text from images,
    and returns it as a list of dictionaries, one for each sheet.

    Args:
        xlsx_path (str): The path to the XLSX file.
        tesseract_cmd (str, optional): The command or path for the Tesseract executable.
                                       Defaults to None, which uses the system's PATH.

    Returns:
        List[Dict[str, Any]]: A list of dictionaries, where each dictionary represents a sheet
                              and has the format: {'page_num': int, 'content': str}.
                              'page_num' is the 1-based sheet number.
                              'content' is a formatted string containing the sheet name,
                              table rows, and image OCR text.
    """
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    wb = load_workbook(xlsx_path, data_only=True)
    processed_sheets = []

    # Enumerate through sheets to get a 1-based page number
    for i, sheet in enumerate(wb.worksheets, start=1):
        sheet_content_lines = []

        # Add the sheet name as a header
        sheet_content_lines.append(f"### Sheet: {sheet.title}")

        # Extract table data
        for row in sheet.iter_rows(max_row=sheet.max_row, values_only=True):
            if all(cell is None for cell in row):
                continue  # Skip completely empty rows
            
            row_data = [str(cell).strip() if cell is not None else "" for cell in row]
            row_content = ",".join(row_data)
            sheet_content_lines.append(f"- {row_content}")

        # Extract images from the sheet
        if hasattr(sheet, '_images'):
            for img in sheet._images:
                image_content = ""
                try:
                    if hasattr(img, '_data'):  # Check if it's a real OpenPyXL Image
                        image_data = img._data()
                    elif hasattr(img, '_ref'): # Skip cell ref-only images
                        continue
                    else:
                        continue

                    pil_img = Image.open(BytesIO(image_data))
                    ocr_text = pytesseract.image_to_string(pil_img).strip()
                    
                    image_content = ocr_text if ocr_text else "[No OCR text detected]"

                except Exception as e:
                    image_content = f"[OCR failed: {str(e)}]"
                
                sheet_content_lines.append(f"[Image OCR Content] {image_content}")

        # Combine all content for the sheet into a single string
        final_sheet_content = "\n".join(sheet_content_lines)

        # Create the dictionary in the desired format
        sheet_dict = {
            "page_num": i,
            "content": final_sheet_content
        }
        
        processed_sheets.append(sheet_dict)

    return processed_sheets
